from bs4 import BeautifulSoup
from datetime import datetime
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import random
import requests
import uuid


def scrape(s_url, url, products_scrape):
    """
    Scrapes product data from the given search URL and saves the page locally.

    Args:
        s_url (str): The search URL to scrape.
        url (str): Main domain URL.
        products_scrape (int, optional): Number of products to scrape.

    Returns:
        bool: True if scraping is successful, False otherwise.
    """
    print(f"🔍 Scraping {products_scrape} {'products' if products_scrape > 1 else 'product'} from {s_url}...\n")

    try:
        # Generate a unique filename for storing the response
        file_name = f"{uuid.uuid4()}.html"
        file_path = os.path.join("temp_files", file_name)

        # Send GET request to the URL
        response = requests.get(s_url, timeout=10)
        response.raise_for_status()  # Raise exception for HTTP errors

        if response.status_code == 200:
            # Save response content to a temporary file
            with open(file_path, "wb") as f:
                f.write(response.content)

            # Call function to extract product details
            result = scrape_products(file_path, products_scrape, url)
            return result
        else:
            print(f"❌ [ERROR] Failed to scrape {s_url}, Status Code: {response.status_code}\n")
            return None
    except requests.ConnectionError:
        print(f"❌ [ERROR] Connection Error: Failed to connect to {s_url}\n")
    except requests.ConnectTimeout:
        print(f"⏳ [ERROR] Timeout Error: {s_url} took too long to respond\n")
    except requests.HTTPError as http_err:
        print(f"❌ [ERROR] HTTP Error: {http_err}\n")
    except requests.RequestException as req_err:
        print(f"❌ [ERROR] Request Error: {req_err}\n")
    except Exception as err:
        print(f"❌ [ERROR] Unexpected Error: {err}\n")
    return None

def scrape_products(file_path, products_scrape, url):
    """
    Extracts product information from the stored HTML file and saves it in Excel.

    Args:
        file_path (str): Path to the stored HTML file.
        products_scrape (int): Number of products to extract.
    """
    # Define HTML class names for product attributes
    products_container_class = "gdgoEp"
    products_class = "cPHDOP"
    # Vertical prodcuts rows
    products_name_class = "KzDlHZ"
    products_price_class = "Nx9bqj"
    products_image_class = "DByuf4"
    products_rating_class = "XQDdHH"
    products_desc_class = "_6NESgJ"

    # Read the saved HTML file
    with open(file_path, "r", encoding="utf-8") as f:
        content = f.read()

    soup = BeautifulSoup(content, "html.parser")

    try:
        # Locate the product container
        products_container = soup.find_all(class_=products_container_class)
        if not products_container:
            # Clean up the temporary file after scraping
            clean(file_path)
            return 0

        products_container = products_container[-1]

        products = products_container.find_all(class_=products_class)

        # Remove unnecessary elements and limit to requested count
        products = products[1:len(products) - 2][:products_scrape]

        # Determine product listing layout: vertical (one per row) or grid (multiple per row)
        is_vertical = True  # Default assumption: products are listed one per row
        product_row = products[0]  # Check the first product row

        # Count the number of product elements in the row to determine layout type
        product_row_products = len(list(product_row.find(class_="_75nlfW").children))
        if product_row_products > 1:
            is_vertical = False  # If multiple products exist in one row, it's a grid layout

        horizontel_row_products = []
        if not is_vertical:
            # Classes for horizontel prodcuts rows
            products_price_class = "Nx9bqj"
            for prod in products:
                horizontel_row_products.extend(list(prod.find(class_="_75nlfW").children))

            products = horizontel_row_products[:products_scrape]

        scrapped_products = []

        # Extract details for each product
        for i, product in enumerate(products):
            product_link = product.find("a").attrs.get("href") if product.find("a") else ""
            if not is_vertical:
                a_tags = product.find_all("a")
                img_tags = product.find_all("img")
                product_image = img_tags[0].attrs.get("src")
                for a_tag in a_tags:
                    title = a_tag.attrs.get("title")
                    if title:
                        product_name = title
                        break
            else:
                product_name = product.find(class_=products_name_class).text if product.find(class_=products_name_class) else ""
                product_image = product.find(class_=products_image_class).attrs.get("src") if product.find(class_=products_image_class) else ""

            product_price = product.find(class_=products_price_class).text if product.find(class_=products_price_class) else ""

            if (product_name == "" and product_price == "" and product_link == "") or product_link == "":
                continue

            product_data = {
                "SNo": i + 1,
                "Name": product_name,
                "Image": product_image,
                "Price": product_price,
                "Link": url + product_link,
            }

            if is_vertical:  # if single product eixts in a row, get desc and rating
                # Extract description list
                product_desc_container = product.find(class_=products_desc_class)
                product_desc_lis = product_desc_container.find_all("li")
                product_desc = [li.text for li in product_desc_lis]

                product_rating = product.find(class_=products_rating_class).text if product.find(class_=products_rating_class) else ""

                product_data["Description"] = product_desc
                product_data["Rating"] = product_rating

            scrapped_products.append(product_data)
    except AttributeError as attr_err:
        print(f"❌ Attribute Error: {attr_err}\n")
        if not scrapped_products:
            return None
    except Exception as err:
        print(f"❌ Unexpected Error: {err}\n")
        return None

    # Clean up the temporary file after scraping
    clean(file_path)

    # Save extracted products to an Excel file
    # save_to_excel(scrapped_products)
    print(f"✅ {len(scrapped_products)} {'products' if products_scrape > 1 else 'product'} scrapped successfully.\n")
    return scrapped_products

def save_to_excel(products):
    """
    📊 Saves the scraped product data into an Excel file with improved formatting.

    Features:
    ✅ Auto-adjusts column widths based on content.
    ✅ Bold headers for better readability.
    ✅ Center-aligns the headers.
    ✅ Description column wrapped to avoid excessive width.

    Args:
        products (list of dict): List containing product details.
    """
    # Convert 'Description' list to a formatted string
    for product in products:
        if "Description" not in product.keys():
            break
        product["Description"] = "\n".join(product["Description"])

    df = pd.DataFrame(products)

    # Generate a timestamped filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"scraped_products_{timestamp}.xlsx"

    # Save the DataFrame to Excel
    df.to_excel(file_name, index=False, engine="openpyxl")

    # Load the workbook to format it
    wb = load_workbook(file_name)
    ws = wb.active

    # Apply formatting
    for col_num, col in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)

        # Determine max content width
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = max_length + 2  # Extra padding
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)  # Limit max width

    # Style headers
    header_font = Font(bold=True)
    for cell in ws[1]:  
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Wrap text in the "Description" column
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ws.max_column, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Save the formatted file
    wb.save(file_name)
    print(f"📁 Data successfully saved and formatted in {file_name} ✅\n")

def clean(file_path):
    """
    🗑️ Cleans up temporary files after scraping.

    Args:
        file_path (str): The path of the temporary file to delete.
    """
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
            print(f"🧹 Successfully deleted temporary file: {file_path}\n")
        else:
            print(f"⚠️  File not found: {file_path}\n")
    except Exception as e:
        print(f"❌ Error while deleting file {file_path}: {e}\n")

def start_flipkart_scrapper(search_key):
    """
    Start scrapping
    """
    # Ensure the 'temp_files' directory exists
    os.makedirs("temp_files", exist_ok=True)

    # Generate a random number between 1 - 50
    products_scrape = random.randint(1, 50)

    url = "https://www.flipkart.com"

    # Construct search URL
    search_url = f"{url}/search?q={search_key}"

    # Call the scrape function with the provided number of products
    result = scrape(search_url, url, products_scrape)

    return result


if __name__ == "__main__":
    # Ensure the 'temp_files' directory exists
    os.makedirs("temp_files", exist_ok=True)

    url = "https://www.flipkart.com"

    # Get search query from user
    search_key = input("🔎 Enter a search key:\n")
    if not search_key:
        print("❌ No search key provided!\n")
        quit()
    elif search_key.isdigit():
        print("❌ Please enter a valid search term (string)!\n")
        quit()

    # Get the number of products to scrape (optional)
    products_scrape = input("🛍️  How many products do you want to scrape? (Optional):\n")
    if products_scrape:
        if not products_scrape.isdigit():
            print("❌ Please enter a valid integer for product count!\n")
            quit()

        products_scrape = int(products_scrape)

        if products_scrape < 1:
            print("❌ Please enter a number greater than 0!\n")
            quit()

        print(f"🔢 You have chosen to scrape {products_scrape} products.\n")
    else:
        products_scrape = 10  # Default value if the user doesn't enter anything
        print("⚙️  No input provided. Scraping default 10 products.\n")

    # Construct search URL
    search_url = f"{url}/search?q={search_key}"

    # Call the scrape function with the provided number of products
    result = scrape(search_url, url, products_scrape)

    # Display final success/failure message
    if result:
        print("🎉 Scraping successful!\n")
    elif result == 0:
        print("⚠️  Products not found\n")
    else:
        print("❌ Scraping failed.\n")
