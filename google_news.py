from bs4 import BeautifulSoup
from datetime import datetime
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import random
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import time
import uuid
from webdriver_manager.chrome import ChromeDriverManager


def get_html_selenium(url, scroll_pause=0.5, scroll_step=300):
    """
    Fetches the full HTML source of a webpage using headless Selenium with smooth scrolling.

    Args:
        url (str): The target URL.
        scroll_pause (float): Time to pause (in seconds) between scrolls.
        scroll_step (int): Number of pixels to scroll per step.
    """
    options = Options()
    options.add_argument("--headless")  # Run in background
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36")

    # Setup WebDriver
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    
    # Open URL
    driver.get(url)
    
    # Wait for JavaScript content to load (Optional)
    driver.implicitly_wait(5)

    # Get total page height
    last_height = driver.execute_script("return document.body.scrollHeight")

    while True:
        for _ in range(0, last_height, scroll_step):
            driver.execute_script(f"window.scrollBy(0, {scroll_step});")
            time.sleep(scroll_pause)  # Pause for data to load

        # Allow time for new content to load
        time.sleep(1)

        # Calculate new scroll height after scrolling
        new_height = driver.execute_script("return document.body.scrollHeight")

        # Break if no new content is loaded
        if new_height == last_height:
            break
        last_height = new_height

    # Get full page source
    page_source = driver.page_source
    
    # Close driver
    driver.quit()
    
    return page_source

def scrape(s_url, url, news_scrape):
    """
    Scrapes news data from the given search URL and saves the page locally.

    Args:
        s_url (str): The search URL to scrape.
        url (str): Main domain URL.
        news_scrape (int, optional): Number of newss to scrape.

    Returns:
        bool: True if scraping is successful, False otherwise.
    """
    print(f"🔍 Scraping {news_scrape} {'newss' if news_scrape > 1 else 'news'} from {s_url}...\n")

    try:
        # Generate a unique filename for storing the response
        file_name = f"{uuid.uuid4()}.html"
        file_path = os.path.join("temp_files", file_name)

        # Get html content to selenium
        html_content = get_html_selenium(s_url)
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

        # Call function to extract news details
        result = scrape_newss(file_path, news_scrape, url)
        return result
    except Exception as err:
        print(f"❌ [ERROR] Unexpected Error: {err}\n")
    return None

def scrape_newss(file_path, news_scrape, url):
    """
    Extracts news information from the stored HTML file and saves it in Excel.

    Args:
        file_path (str): Path to the stored HTML file.
        news_scrape (int): Number of newss to extract.
    """
    # Define HTML class names for news attributes
    newss_container_class = "D9SJMe"

    # Read the saved HTML file
    with open(file_path, "r", encoding="utf-8") as f:
        content = f.read()

    soup = BeautifulSoup(content, "html.parser")

    try:
        # Locate the news container
        newss_container = soup.find(class_=newss_container_class)

        newss = newss_container.children
        newss = list(newss)
        print('Total:', len(newss), '\n')
        if not newss:
            # Clean up the temporary file after scraping
            clean(file_path)
            return 0

        # Limit to requested count
        newss = newss[5:news_scrape + 5]
        print('Total:', len(newss), '\n')

        scrapped_newss = []

        # Extract details for each news
        for i, news in enumerate(newss):
            news_provider = news.find(class_="zC7z7b").attrs.get("src") if news.find(class_="zC7z7b") else ""
            news_title = news.find(class_="JtKRv").text if news.find(class_="JtKRv") else ""
            news_image = news.find(class_="Quavad").attrs.get("src") if news.find(class_="Quavad") else ""
            news_time = news.find(class_="hvbAAd").attrs.get("datetime") if news.find(class_="hvbAAd") else ""
            news_link = news.find(class_="WwrzSb").attrs.get('href') if news.find(class_="WwrzSb") else ""

            if (news_title == "" and news_link == "" and news_provider == "") or news_link == "":
                continue

            news_data = {
                "SNo": i + 1,
                "Provider": news_provider,
                "Title": news_title,
                "Image": url + news_image,
                "Time": datetime.strptime(news_time, "%Y-%m-%dT%H:%M:%SZ").strftime("%d %b %Y") if news_time else "",
                "Link": url + news_link,
            }

            scrapped_newss.append(news_data)

    except AttributeError as attr_err:
        print(f"❌ Attribute Error: {attr_err}\n")
        if not scrapped_newss:
            return None
    except Exception as err:
        print(f"❌ Unexpected Error: {err}\n")
        return None

    # Clean up the temporary file after scraping
    clean(file_path)

    # Save extracted newss to an Excel file
    # save_to_excel(scrapped_newss)
    print(f"✅ {len(scrapped_newss)} {'newss' if news_scrape > 1 else 'news'} scrapped successfully.\n")
    return scrapped_newss

def save_to_excel(newss):
    """
    📊 Saves the scraped news data into an Excel file with improved formatting.

    Features:
    ✅ Auto-adjusts column widths based on content.
    ✅ Bold headers for better readability.
    ✅ Center-aligns the headers.

    Args:
        newss (list of dict): List containing news details.
    """

    df = pd.DataFrame(newss)

    # Generate a timestamped filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"scraped_newss_{timestamp}.xlsx"

    # Save the DataFrame to Excel
    df.to_excel(file_name, index=False, engine="openpyxl")

    # Load the workbook to format it
    wb = load_workbook(file_name)
    ws = wb.active

    # Apply formatting
    for col_num, col in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)

        # Determine max content width
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = max_length + 2  # Extra padding
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)  # Limit max width

    # Style headers
    header_font = Font(bold=True)
    for cell in ws[1]:  
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Wrap text in the "Description" column
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ws.max_column, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Save the formatted file
    wb.save(file_name)
    print(f"📁 Data successfully saved and formatted in {file_name} ✅\n")

def clean(file_path):
    """
    🗑️ Cleans up temporary files after scraping.

    Args:
        file_path (str): The path of the temporary file to delete.
    """
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
            print(f"🧹 Successfully deleted temporary file: {file_path}\n")
        else:
            print(f"⚠️  File not found: {file_path}\n")
    except Exception as e:
        print(f"❌ Error while deleting file {file_path}: {e}\n")

def start_g_news_scrapper(search_key):
    """
    Start scrapping
    """
    # Ensure the 'temp_files' directory exists
    os.makedirs("temp_files", exist_ok=True)

    # Generate a random number between 1 - 50
    news_scrape = random.randint(1, 50)

    url = "https://news.google.com"

    # Construct search URL
    search_url = f"{url}/search?q={search_key}"

    # Call the scrape function with the provided number of products
    result = scrape(search_url, url, news_scrape)

    return result


if __name__ == "__main__":
    # Ensure the 'temp_files' directory exists
    os.makedirs("temp_files", exist_ok=True)

    url = "https://news.google.com"

    # Get search query from user
    search_key = input("🔎 Enter a search key:\n")
    if not search_key:
        print("❌ No search key provided!\n")
        quit()
    elif search_key.isdigit():
        print("❌ Please enter a valid search term (string)!\n")
        quit()

    # Get the number of news to scrape (optional)
    news_scrape = input("🛍️  How many news do you want to scrape? (Optional):\n")
    if news_scrape:
        if not news_scrape.isdigit():
            print("❌ Please enter a valid integer for news count!\n")
            quit()

        news_scrape = int(news_scrape)

        if news_scrape < 1:
            print("❌ Please enter a number greater than 0!\n")
            quit()

        print(f"🔢 You have chosen to scrape {news_scrape} news.\n")
    else:
        news_scrape = 10  # Default value if the user doesn't enter anything
        print("⚙️  No input provided. Scraping default 10 news.\n")

    # Construct search URL
    search_url = f"{url}/search?q={search_key}"

    # Call the scrape function with the provided number of news
    result = scrape(search_url, url, news_scrape)

    # Display final success/failure message
    if result:
        print("🎉 Scraping successful!\n")
    elif result == 0:
        print("⚠️  News not found\n")
    else:
        print("❌ Scraping failed.\n")
